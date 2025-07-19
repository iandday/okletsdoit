import logging
from typing import Any, Mapping
from django.shortcuts import redirect, render, get_object_or_404
from django.db.models import Sum, Case, When, Value, DecimalField, F
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import HttpRequest, JsonResponse, HttpResponse
from django.contrib import messages
from django.utils.text import slugify
from django.urls import reverse
import polars as pl
import openpyxl
from io import BytesIO
from datetime import datetime
from decimal import Decimal, InvalidOperation
from expenses.forms import CategoryForm, ExpenseForm
from list.models import ListEntry
from .models import Category, Expense


logger = logging.getLogger(__name__)


@login_required
def summary(request):
    # Get all categories with their expense totals and calculations
    categories = (
        Category.objects.filter(is_deleted=False)
        .annotate(total_estimated=Sum("expense__estimated_amount"), total_actual=Sum("expense__actual_amount"))
        .annotate(
            # Calculate percentage
            percentage=Case(
                When(
                    total_estimated__gt=0,
                    total_actual__isnull=False,
                    then=(F("total_actual") / F("total_estimated")) * 100,
                ),
                default=Value(None),
                output_field=DecimalField(max_digits=5, decimal_places=1),
            ),
        )
        .filter(total_estimated__isnull=False)
        .order_by("name")
    )

    # Calculate overall totals
    overall_estimated = (
        Expense.objects.filter(is_deleted=False, estimated_amount__isnull=False).aggregate(
            total=Sum("estimated_amount")
        )["total"]
        or 0
    )

    overall_actual = (
        Expense.objects.filter(is_deleted=False, actual_amount__isnull=False).aggregate(total=Sum("actual_amount"))[
            "total"
        ]
        or 0
    )

    # Add calculated fields to each category
    categories_with_calcs = []
    estimate_labels = []
    actual_labels = []
    estimate_data = []
    actual_data = []
    estimate_slugs = []
    actual_slugs = []

    for category in categories:
        category_data = {
            "category": category,
            "total_estimated": category.total_estimated,
            "total_actual": category.total_actual,
            "percentage": category.percentage,
        }
        categories_with_calcs.append(category_data)

        # Fix chart data logic
        if category.total_estimated and category.total_estimated > 0:
            estimate_labels.append(category.name)
            estimate_data.append(float(category.total_estimated))
            estimate_slugs.append(category.slug)

        if category.total_actual and category.total_actual > 0:
            actual_labels.append(category.name)
            actual_data.append(float(category.total_actual))
            actual_slugs.append(category.slug)

    context = {
        "categories": categories_with_calcs,
        "overall_estimated": overall_estimated,
        "overall_actual": overall_actual,
        "estimate_labels": estimate_labels,
        "estimate_data": estimate_data,
        "actual_labels": actual_labels,
        "actual_data": actual_data,
        "estimate_slugs": estimate_slugs,
        "actual_slugs": actual_slugs,
    }

    return render(request, "expenses/summary.html", context)


@login_required
def create(request):
    if request.method == "POST":
        form = ExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.created_by = request.user
            expense.save()
    return redirect("expenses:list")

    return render(request, "expenses/create.html")


@login_required
def list(request: HttpRequest):
    # check if category is in parameters
    category_slug = request.GET.get("category")
    context = {}
    context["form"] = ExpenseForm()
    if category_slug:
        category = Category.objects.filter(slug=category_slug, is_deleted=False)
        if not category:
            return redirect("expenses:category_list")
        context["category"] = category  # type: ignore[assignment]
    context["expenses"] = Expense.objects.filter(is_deleted=False).order_by("-date")  # type: ignore[assignment]

    return render(request, "expenses/expense_list.html", context)


@login_required
def expense_edit(request, slug: str):
    """Edit an existing expense"""
    expense = get_object_or_404(Expense, slug=slug, is_deleted=False)

    if request.method == "POST":
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.updated_by = request.user
            expense.save()
            messages.success(request, f"Expense '{expense.item}' updated successfully.")
            return redirect("expenses:detail", slug=expense.slug)
        else:
            # Add field-specific error messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            messages.error(request, "Please correct the errors below.")
    else:
        form = ExpenseForm(instance=expense)

    return render(request, "expenses/expense_form.html", {"form": form, "expense": expense})


@login_required
def expense_data(request):
    """JSON endpoint for DataTables"""
    category_slug = request.GET.get("category")

    if category_slug:
        expenses = Expense.objects.filter(category__slug=category_slug, is_deleted=False).select_related("category")
    else:
        expenses = Expense.objects.filter(is_deleted=False).select_related("category")

    data = []
    for expense in expenses:
        if expense.list_entries.exists():
            est_amount = float(expense.calculated_estimated_amount())
            actual_amount = float(expense.calculated_actual_amount())
            item = f'<a href="{reverse("expenses:detail", args=[expense.slug])}" class="link link-primary"><span class="italic">{expense.item}</span></a>'
        else:
            est_amount = float(expense.estimated_amount) if expense.estimated_amount else 0
            actual_amount = float(expense.actual_amount) if expense.actual_amount else 0
            item = f'<a href="{reverse("expenses:detail", args=[expense.slug])}" class="link link-primary">{expense.item}</a>'

        # add link to the associated list in the drop-down menu
        menu_content = """
            <div class="dropdown dropdown-end">
                <div tabindex="0" role="button" class="btn btn-ghost btn-sm">
                    <svg xmlns="http://www.w3.org/2000/svg" class="h-4 w-4" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M12 5v.01M12 12v.01M12 19v.01" />
                    </svg>
                </div>
                <ul tabindex="0" class="dropdown-content menu bg-base-100 rounded-box z-[1] w-32 p-2 shadow">
        """
        if expense.list_entries.exists():
            menu_content += f'<li><a href="{reverse("list:detail", args=[expense.list_entries.first().list.slug])}" class="link link-primary">List</a></li>'
        menu_content += f"""
                    <li><a href="{reverse("expenses:expense_edit", args=[expense.slug])}" class="text-sm">Edit</a></li>
                </ul>
            </div>
        """

        data.append(
            {
                "id": str(expense.id),
                "item": item,
                "category": expense.category.name if expense.category else "Uncategorized",
                "date": expense.date.strftime("%Y-%m-%d") if expense.date else None,
                "estimated_amount": est_amount,
                "actual_amount": actual_amount,
                "slug": expense.slug,
                "menu_content": menu_content,
            }
        )
    return JsonResponse({"data": data})


@login_required
def expense_import(request):
    """Import expenses from Excel file"""
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please select an Excel file to upload.")
            return redirect("expenses:summary")

        if not excel_file.name.endswith((".xlsx", ".xls")):
            messages.error(request, "Please upload a valid Excel file (.xlsx or .xls).")
            return redirect("expenses:summary")

        try:
            # Read the Excel file
            df = pl.read_excel(BytesIO(excel_file.read()))

            # Validate required columns
            required_columns = ["item"]
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
                return redirect("expenses:summary")

            # Import expenses
            created_count = 0
            updated_count = 0
            error_count = 0
            created_categories = 0

            for row in df.to_dicts():
                try:
                    item = str(row.get("item", "")).strip()
                    if not item or item.lower() == "nan":
                        error_count += 1
                        continue

                    # Handle category - create if doesn't exist
                    category = None
                    category_name = str(row.get("category", "")).strip()
                    if category_name and category_name.lower() != "nan":
                        category, created = Category.objects.get_or_create(
                            name=category_name,
                            defaults={
                                "slug": slugify(category_name),
                                "created_by": request.user,
                                "is_deleted": False,
                            },
                        )
                        if created:
                            created_categories += 1

                    # Parse date
                    expense_date = None
                    date_str = str(row.get("date", "")).strip()
                    if date_str and date_str.lower() != "nan":
                        try:
                            expense_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                        except ValueError:
                            try:
                                expense_date = datetime.strptime(date_str, "%m/%d/%Y").date()
                            except ValueError:
                                expense_date = None

                    # Parse amounts
                    estimated_amount = None
                    actual_amount = None
                    quantity = 1

                    estimated_str = str(row.get("estimated_amount", "")).strip()
                    if estimated_str and estimated_str.lower() != "nan":
                        try:
                            estimated_amount = Decimal(str(estimated_str).replace("$", "").replace(",", ""))
                        except (InvalidOperation, ValueError):
                            estimated_amount = None

                    actual_str = str(row.get("actual_amount", "")).strip()
                    if actual_str and actual_str.lower() != "nan":
                        try:
                            actual_amount = Decimal(str(actual_str).replace("$", "").replace(",", ""))
                        except (InvalidOperation, ValueError):
                            actual_amount = None

                    quantity_str = str(row.get("quantity", "")).strip()
                    if quantity_str and quantity_str.lower() != "nan":
                        try:
                            quantity = int(float(quantity_str))
                        except (ValueError, TypeError):
                            quantity = 1

                    # Check if expense already exists
                    existing_expense = None
                    if category:
                        existing_expense = Expense.objects.filter(
                            item__iexact=item, category=category, is_deleted=False
                        ).first()
                    else:
                        existing_expense = Expense.objects.filter(
                            item__iexact=item, category__isnull=True, is_deleted=False
                        ).first()

                    expense_data = {
                        "item": item,
                        "description": str(row.get("description", "")).strip()
                        if row.get("description", None) is not None
                        else "",
                        "date": expense_date,
                        "category": category,
                        "quantity": quantity,
                        "estimated_amount": estimated_amount,
                        "actual_amount": actual_amount,
                        "url": str(row.get("url", "")).strip() if row.get("url", None) is not None else "",
                    }

                    # Clean empty strings
                    for key, value in expense_data.items():
                        if isinstance(value, str) and (value.lower() == "nan" or not value):
                            expense_data[key] = "" if key not in ["item"] else expense_data[key]

                    if existing_expense:
                        # Update existing expense
                        for key, value in expense_data.items():
                            if value is not None and value != "":  # Only update non-empty values
                                setattr(existing_expense, key, value)
                        existing_expense.updated_by = request.user
                        existing_expense.save()
                        updated_count += 1
                    else:
                        # Create new expense
                        expense = Expense(**expense_data)
                        expense.created_by = request.user
                        expense.save()
                        created_count += 1

                except Exception as e:
                    logger.error(e)
                    error_count += 1
                    continue

            # Success message
            message_parts = []
            if created_count > 0:
                message_parts.append(f"{created_count} expense{'s' if created_count != 1 else ''} created")
            if updated_count > 0:
                message_parts.append(f"{updated_count} expense{'s' if updated_count != 1 else ''} updated")
            if created_categories > 0:
                message_parts.append(
                    f"{created_categories} new categor{'ies' if created_categories != 1 else 'y'} created"
                )

            if message_parts:
                messages.success(request, f"Import completed: {', '.join(message_parts)}.")

            if error_count > 0:
                messages.warning(request, f"{error_count} row{'s' if error_count != 1 else ''} skipped due to errors.")

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")

    return redirect("expenses:summary")


@login_required
def detail(request, slug: str):
    """View for expense detail"""
    expense = get_object_or_404(Expense, slug=slug, is_deleted=False)
    context = {
        "expense": expense,
        "form": ExpenseForm(instance=expense),
        "list_entries": expense.list_entries.filter(is_deleted=False).order_by("order"),  # type: ignore
    }
    return render(request, "expenses/expense_detail.html", context)


@login_required
def template_download(request):
    """Download Excel template for expense import"""
    # Create a workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses Template"

    # Define headers
    headers = ["item", "description", "date", "category", "quantity", "estimated_amount", "actual_amount", "url"]

    # Add headers to the first row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)

    # Add sample data
    sample_data = [
        [
            "Wedding Photography",
            "Professional photographer for ceremony and reception",
            "2024-08-15",
            "Photography",
            1,
            2500.00,
            2800.00,
            "https://example.com/wedding-photography",
        ],
        [
            "Wedding Cake",
            "3-tier vanilla cake with custom decorations",
            "2024-08-15",
            "Catering",
            1,
            800.00,
            750.00,
            "https://example.com/wedding-cake",
        ],
        [
            "Bridal Dress",
            "Designer wedding dress with alterations",
            "2024-06-01",
            "Attire",
            1,
            1500.00,
            1650.00,
            "https://example.com/bridal-dress",
        ],
        [
            "Flower Arrangements",
            "Bridal bouquet and centerpieces",
            "2024-08-15",
            "Flowers",
            8,
            150.00,
            140.00,
            "https://example.com/flower-arrangements",
        ],
        [
            "Wedding Venue",
            "Reception hall rental for 6 hours",
            "2024-08-15",
            "Venue",
            1,
            3000.00,
            "",
            "https://example.com/wedding-venue",
        ],
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = HttpResponse(
        output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="expenses_template.xlsx"'

    return response


@login_required
def category_detail(request, slug: str):
    category = get_object_or_404(Category, slug=slug, is_deleted=False)
    expenses = Expense.objects.filter(category=category, is_deleted=False)
    context = {
        "category": category,
        "expenses": expenses,
        "total_estimated": expenses.aggregate(total=Sum("estimated_amount", filter=Q(estimated_amount__isnull=False)))[
            "total"
        ]
        or Decimal("0.00"),
        "total_actual": expenses.aggregate(total=Sum("actual_amount", filter=Q(actual_amount__isnull=False)))["total"]
        or Decimal("0.00"),
    }
    return render(request, "expenses/category_detail.html", context)


@login_required
def category_list(request):
    categories = (
        Category.objects.filter(is_deleted=False)
        .prefetch_related("expense_set")
        .annotate(
            total_expenses=Count("expense", filter=Q(expense__is_deleted=False)),
            expenses_with_actual=Count(
                "expense", filter=Q(expense__actual_amount__isnull=False, expense__is_deleted=False)
            ),
            expenses_with_estimated=Count(
                "expense", filter=Q(expense__estimated_amount__isnull=False, expense__is_deleted=False)
            ),
        )
        .order_by("name")
    )

    form = CategoryForm()

    category_forms = {}
    for category in categories:
        category_forms[category.id] = CategoryForm(instance=category)

    return render(
        request,
        "expenses/category_list.html",
        {"categories": categories, "form": form, "category_forms": category_forms},
    )


def category_create(request):
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            category: Category = form.save(commit=False)
            category.created_by = request.user
            category.save()
    return redirect("expenses:category_list")


def category_edit(request, slug: str):
    category = Category.objects.get(slug=slug, is_deleted=False)
    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            category = form.save(commit=False)
            category.updated_by = request.user
            category.save()
    return redirect("expenses:category_list")


def category_delete(request, slug: str):
    category = Category.objects.get(slug=slug, is_deleted=False)
    if request.method == "POST":
        category.is_deleted = True
        category.save()
    return redirect("expenses:category_list")
