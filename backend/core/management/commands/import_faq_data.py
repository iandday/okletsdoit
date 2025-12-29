"""
Management command to import FAQ data from Excel file.

Usage:
    python manage.py import_faq_data --file faq_data.xlsx
    python manage.py import_faq_data --generate-template faq_template.xlsx

Excel File Format:

    The Excel file should contain 2 sheets:

    1. Questions:
       Columns: category_name, category_order, category_published, question, answer,
                question_order, icon, question_published, url_1, url_1_text, url_2, url_2_text, url_3, url_3_text

    2. Tips:
       Columns: category_name, content, tip_order, tip_published
"""

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from core.models import QuestionCategory, Question, QuestionURL, Tips
from users.models import User

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    raise CommandError("openpyxl is required. Install it with: pip install openpyxl")


class Command(BaseCommand):
    help = "Import FAQ data from Excel file or generate an Excel template"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            help="Path to Excel file to import",
        )
        parser.add_argument(
            "--generate-template",
            type=str,
            help="Generate an empty Excel template file at the specified path",
        )
        parser.add_argument(
            "--user",
            type=str,
            default="admin",
            help="Username of the user to set as creator (default: admin)",
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Clear existing FAQ data before importing",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Perform a dry run without saving to database",
        )

    def handle(self, *args, **options):
        # Generate template if requested
        if options["generate_template"]:
            self._generate_template(options["generate_template"])
            return

        # Import from file
        if not options["file"]:
            raise CommandError("Either --file or --generate-template must be provided")

        # Get the user
        try:
            user = User.objects.get(email=options["user"])
        except User.DoesNotExist:
            raise CommandError(f"User '{options['user']}' does not exist")

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("DRY RUN MODE - No changes will be saved"))

        try:
            with transaction.atomic():
                # Clear existing data if requested
                if options["clear"]:
                    if options["dry_run"]:
                        self.stdout.write("Would delete existing FAQ data")
                    else:
                        self._clear_data()

                # Load Excel file
                try:
                    wb = openpyxl.load_workbook(options["file"])
                except FileNotFoundError:
                    raise CommandError(f"File not found: {options['file']}")
                except Exception as e:
                    raise CommandError(f"Error loading Excel file: {str(e)}")

                # Import data from sheets
                questions_created, categories_created, urls_created = self._import_questions(
                    wb, user, options["dry_run"]
                )
                tips_created = self._import_tips(wb, user, options["dry_run"])

                # Rollback if dry run
                if options["dry_run"]:
                    transaction.set_rollback(True)
                    self.stdout.write(self.style.WARNING("\nDry run complete - no changes saved"))
                else:
                    self.stdout.write(self.style.SUCCESS("\n✓ Import completed successfully"))

                # Summary
                self.stdout.write(f"\nSummary:")
                if categories_created > 0:
                    self.stdout.write(f"  Categories: {categories_created}")
                if questions_created > 0:
                    self.stdout.write(f"  Questions: {questions_created}")
                if urls_created > 0:
                    self.stdout.write(f"  URLs: {urls_created}")
                if tips_created > 0:
                    self.stdout.write(f"  Tips: {tips_created}")

        except Exception as e:
            raise CommandError(f"Import failed: {str(e)}")

    def _generate_template(self, filepath):
        """Generate an empty Excel template with example data"""
        self.stdout.write(f"Generating template Excel file: {filepath}")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="67260B", end_color="67260B", fill_type="solid")

        # Create Questions sheet
        ws_questions = wb.create_sheet("Questions")
        ws_questions.append(
            [
                "category_name",
                "category_order",
                "category_published",
                "question",
                "answer",
                "question_order",
                "icon",
                "question_published",
                "url_1",
                "url_1_text",
                "url_2",
                "url_2_text",
                "url_3",
                "url_3_text",
            ]
        )
        for cell in ws_questions[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Add example data
        ws_questions.append(
            [
                "Getting Started",
                1,
                "true",
                "When is the wedding?",
                "The wedding will take place on June 15, 2026.",
                1,
                "calendar",
                "true",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        ws_questions.append(
            [
                "Getting Started",
                1,
                "true",
                "Where is the venue?",
                "The ceremony and reception will be held at Riverside Manor, 123 Main Street.",
                2,
                "map-pin",
                "true",
                "https://maps.google.com/maps?q=Riverside+Manor",
                "View on Google Maps",
                "https://riversidemanor.com",
                "Venue Website",
                "",
                "",
            ]
        )
        ws_questions.append(
            [
                "Venue Details",
                2,
                "true",
                "Is there parking available?",
                "Yes, there is ample free parking on-site for all guests.",
                1,
                "car",
                "true",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )

        # Adjust column widths
        ws_questions.column_dimensions["A"].width = 20  # category_name
        ws_questions.column_dimensions["B"].width = 15  # category_order
        ws_questions.column_dimensions["C"].width = 18  # category_published
        ws_questions.column_dimensions["D"].width = 40  # question
        ws_questions.column_dimensions["E"].width = 60  # answer
        ws_questions.column_dimensions["F"].width = 15  # question_order
        ws_questions.column_dimensions["G"].width = 15  # icon
        ws_questions.column_dimensions["H"].width = 18  # question_published
        ws_questions.column_dimensions["I"].width = 50  # url_1
        ws_questions.column_dimensions["J"].width = 30  # url_1_text
        ws_questions.column_dimensions["K"].width = 50  # url_2
        ws_questions.column_dimensions["L"].width = 30  # url_2_text
        ws_questions.column_dimensions["M"].width = 50  # url_3
        ws_questions.column_dimensions["N"].width = 30  # url_3_text

        # Create Tips sheet
        ws_tips = wb.create_sheet("Tips")
        ws_tips.append(["category_name", "content", "tip_order", "tip_published"])
        for cell in ws_tips[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Add example data
        ws_tips.append(
            [
                "Getting Started",
                "RSVP by the deadline to ensure your spot at the celebration.",
                1,
                "true",
            ]
        )
        ws_tips.append(
            [
                "Getting Started",
                "Book your accommodation early, as rooms fill up quickly during wedding season.",
                2,
                "true",
            ]
        )
        ws_tips.append(
            [
                "Venue Details",
                "The venue has both indoor and outdoor spaces for photos.",
                1,
                "true",
            ]
        )

        # Adjust column widths
        ws_tips.column_dimensions["A"].width = 20  # category_name
        ws_tips.column_dimensions["B"].width = 70  # content
        ws_tips.column_dimensions["C"].width = 12  # tip_order
        ws_tips.column_dimensions["D"].width = 15  # tip_published

        # Save the workbook
        try:
            wb.save(filepath)
            self.stdout.write(self.style.SUCCESS(f"✓ Template created successfully: {filepath}"))
            self.stdout.write("\nThe template contains 2 sheets:")
            self.stdout.write("  1. Questions - Define categories and add FAQ questions with up to 3 URLs per question")
            self.stdout.write("  2. Tips - Add helpful tips that reference categories defined in Questions sheet")
            self.stdout.write("\nEach sheet includes example data that you can modify or delete.")
            self.stdout.write(
                "\nNote: Categories are defined in the Questions sheet. Tips reference these categories by name."
            )
        except Exception as e:
            raise CommandError(f"Error saving template: {str(e)}")

    def _clear_data(self):
        """Clear existing FAQ data"""
        self.stdout.write("Clearing existing FAQ data...")
        QuestionURL.objects.all().delete()
        Question.objects.all().delete()
        Tips.objects.all().delete()
        QuestionCategory.objects.all().delete()
        self.stdout.write(self.style.SUCCESS("✓ Existing data cleared"))

    def _import_questions(self, wb, user, dry_run):
        """Import questions with categories and URLs from Excel sheet"""
        if "Questions" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("  ⚠ No Questions sheet found, skipping"))
            return 0, 0, 0

        self.stdout.write("\nImporting questions and categories...")
        ws = wb["Questions"]
        questions_created = 0
        categories_created = 0
        urls_created = 0
        created_categories = {}

        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[3]:  # Skip rows without category_name or question
                continue

            # Parse row data
            category_name = str(row[0]).strip()
            category_order = int(row[1]) if row[1] else 0
            category_published = str(row[2]).lower() in ["true", "1", "yes"] if row[2] else False
            question_text = str(row[3]).strip()
            answer = str(row[4]).strip() if row[4] else ""
            question_order = int(row[5]) if row[5] else 0
            icon = str(row[6]).strip() if row[6] else "question-circle"
            question_published = str(row[7]).lower() in ["true", "1", "yes"] if row[7] else False

            # Parse URLs (up to 3)
            urls = []
            for i in range(3):
                url_col = 8 + (i * 2)
                url_text_col = 9 + (i * 2)
                if url_col < len(row) and row[url_col]:
                    url = str(row[url_col]).strip()
                    url_text = str(row[url_text_col]).strip() if url_text_col < len(row) and row[url_text_col] else ""
                    urls.append((url, url_text))

            if dry_run:
                if category_name not in created_categories:
                    self.stdout.write(f"  Would create category: {category_name}")
                    created_categories[category_name] = True
                self.stdout.write(f"  Would create question: {question_text[:50]}...")
                for url, url_text in urls:
                    self.stdout.write(f"    Would create URL: {url}")
            else:
                # Create or update category
                if category_name not in created_categories:
                    category, created = QuestionCategory.objects.get_or_create(
                        name=category_name,
                        defaults={
                            "order": category_order,
                            "published": category_published,
                            "created_by": user,
                        },
                    )
                    if created:
                        self.stdout.write(f"  ✓ Created category: {category_name}")
                        categories_created += 1
                    else:
                        # Update existing category
                        category.order = category_order
                        category.published = category_published
                        category.updated_by = user
                        category.save()
                        self.stdout.write(f"  ↻ Updated category: {category_name}")
                    created_categories[category_name] = category
                else:
                    category = created_categories[category_name]

                # Create or update question
                question, created = Question.objects.get_or_create(
                    question=question_text,
                    defaults={
                        "category": category,
                        "answer": answer,
                        "order": question_order,
                        "icon": icon,
                        "published": question_published,
                        "created_by": user,
                    },
                )
                if created:
                    self.stdout.write(f"  ✓ Created question: {question_text[:50]}...")
                    questions_created += 1
                else:
                    # Update existing question
                    question.category = category
                    question.answer = answer
                    question.order = question_order
                    question.icon = icon
                    question.published = question_published
                    question.updated_by = user
                    question.save()
                    self.stdout.write(f"  ↻ Updated question: {question_text[:50]}...")

                # Create or update URLs
                for url, url_text in urls:
                    question_url, created = QuestionURL.objects.get_or_create(
                        url=url,
                        defaults={
                            "question": question,
                            "text": url_text,
                            "created_by": user,
                        },
                    )
                    if created:
                        question.urls.add(question_url)
                        self.stdout.write(f"    ✓ Created URL: {url}")
                        urls_created += 1
                    else:
                        # Update existing URL
                        question_url.question = question
                        question_url.text = url_text
                        question_url.updated_by = user
                        question_url.save()
                        question.urls.add(question_url)
                        self.stdout.write(f"    ↻ Updated URL: {url}")

        return questions_created, categories_created, urls_created

    def _import_tips(self, wb, user, dry_run):
        """Import tips from Excel sheet, referencing existing categories"""
        if "Tips" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("  ⚠ No Tips sheet found, skipping"))
            return 0

        self.stdout.write("\nImporting tips...")
        ws = wb["Tips"]
        tips_created = 0

        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:  # Skip rows without category_name or content
                continue

            # Parse row data
            category_name = str(row[0]).strip()
            content = str(row[1]).strip()
            tip_order = int(row[2]) if row[2] else 0
            tip_published = str(row[3]).lower() in ["true", "1", "yes"] if row[3] else False

            if dry_run:
                self.stdout.write(f"  Would create tip: {content[:50]}...")
            else:
                # Get existing category (created from Questions sheet)
                try:
                    category = QuestionCategory.objects.get(name=category_name)
                except QuestionCategory.DoesNotExist:
                    self.stdout.write(
                        self.style.WARNING(
                            f"  ⚠ Category '{category_name}' not found. Please ensure it's defined in the Questions sheet. Skipping tip."
                        )
                    )
                    continue

                # Create or update tip
                tip, created = Tips.objects.get_or_create(
                    content=content,
                    category=category,
                    defaults={
                        "order": tip_order,
                        "published": tip_published,
                        "created_by": user,
                    },
                )
                if created:
                    self.stdout.write(f"  ✓ Created tip: {content[:50]}...")
                    tips_created += 1
                else:
                    # Update existing tip
                    tip.order = tip_order
                    tip.published = tip_published
                    tip.updated_by = user
                    tip.save()
                    self.stdout.write(f"  ↻ Updated tip: {content[:50]}...")

        return tips_created
