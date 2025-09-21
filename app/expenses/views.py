import logging
from typing import Any, Mapping
from django.shortcuts import redirect, render, get_object_or_404
from django.db.models import Sum, Case, When, Value, DecimalField, F
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import HttpRequest, JsonResponse, HttpResponse, QueryDict
from django.contrib import messages
from django.utils.text import slugify
from django.urls import reverse
import polars as pl
import openpyxl
from io import BytesIO
from datetime import datetime
from decimal import Decimal, InvalidOperation

from attachments.models import Attachment, AttachmentManager
from expenses.forms import CategoryForm, ExpenseForm
from list.models import ListEntry
from .models import Category, Expense
from attachments.forms import AttachmentUploadForm

logger = logging.getLogger(__name__)


def format_row_action_cell(detail_url: str, update_url: str, slug: str) -> str:
    """
    Format a row action cell with links for detail, update, and delete actions.

    Args:
        detail_url (str): URL for the detail view.
        update_url (str): URL for the update view.
        slug (str): object slug.

    Returns:
        str: HTML string with action links.
    """
    return f"""
            <div class="flex flex-row gap-1">
                <a class="btn btn-sm btn-primary" href="{detail_url}">
                    <svg xmlns="http://www.w3.org/2000/svg" class="h-4 w-4" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" 
                        d="M15 12a3 3 0 11-6 0 3 3 0 016 0z" />
                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" 
                        d="M2.458 12C3.732 7.943 7.523 5 12 5c4.478 0 8.268 2.943 9.542 7-1.274 4.057-5.064 7-9.542 7-4.477 0-8.268-2.943-9.542-7z" />
                    </svg> 
                </a>
                <a class="btn btn-sm btn-secondary" href="{update_url}">
                    <svg xmlns="http://www.w3.org/2000/svg" class="h-4 w-4" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" 
                        d="M16.862 3.487l2.651 2.651a1.875 1.875 0 010 2.652L8.288 20.015a4.5 4.5 0 01-1.897 1.13l-3.38.96a.75.75 0 01-.926-.926l.96-3.38a4.5 4.5 0 011.13-1.897L16.862 3.487z" />'
                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" 
                        d="M19.5 7.125L16.862 4.487" />
                    </svg>
                </a>
                <button class="btn btn-sm btn-error delete-btn" data-id="{slug}">
                    <svg xmlns="http://www.w3.org/2000/svg" class="h-4 w-4" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" 
                    d="M3 6h18M9 6V4h6v2m2 0v12a2 2 0 01-2 2H7a2 2 0 01-2-2V6h14z" />
                    </svg> 
                </button>
            </div>
        """


@login_required
def summary(request) -> HttpResponse:
    # Get all categories with their expense totals and calculations
    categories = (
        Category.objects.filter(is_deleted=False)
        .annotate(total_estimated=Sum("expense__estimated_amount"), total_actual=Sum("expense__actual_amount"))
        .annotate(
            # Calculate percentage
            percentage=Case(
                When(
                    total_estimated__gt=0,
                    total_actual__isnull=False,
                    then=(F("total_actual") / F("total_estimated")) * 100,
                ),
                default=Value(None),
                output_field=DecimalField(max_digits=5, decimal_places=1),
            ),
        )
        .filter(total_estimated__isnull=False)
        .order_by("name")
    )

    # Calculate overall totals
    overall_estimated = (
        Expense.objects.filter(is_deleted=False, estimated_amount__isnull=False).aggregate(
            total=Sum("estimated_amount")
        )["total"]
        or 0
    )

    overall_actual = (
        Expense.objects.filter(is_deleted=False, actual_amount__isnull=False).aggregate(total=Sum("actual_amount"))[
            "total"
        ]
        or 0
    )

    # Add calculated fields to each category
    categories_with_calcs = []
    estimate_labels = []
    actual_labels = []
    estimate_data = []
    actual_data = []
    estimate_slugs = []
    actual_slugs = []

    for category in categories:
        category_data = {
            "category": category,
            "total_estimated": category.total_estimated,
            "total_actual": category.total_actual,
            "percentage": category.percentage,
        }
        categories_with_calcs.append(category_data)

        # Fix chart data logic
        if category.total_estimated and category.total_estimated > 0:
            estimate_labels.append(category.name)
            estimate_data.append(float(category.total_estimated))
            estimate_slugs.append(category.slug)

        if category.total_actual and category.total_actual > 0:
            actual_labels.append(category.name)
            actual_data.append(float(category.total_actual))
            actual_slugs.append(category.slug)

    context = {
        "categories": categories_with_calcs,
        "overall_estimated": overall_estimated,
        "overall_actual": overall_actual,
        "estimate_labels": estimate_labels,
        "estimate_data": estimate_data,
        "actual_labels": actual_labels,
        "actual_data": actual_data,
        "estimate_slugs": estimate_slugs,
        "actual_slugs": actual_slugs,
    }

    return render(request, "expenses/summary.html", context)


@login_required
def create(request) -> HttpResponse:
    if request.method == "POST":
        form = ExpenseForm(request.POST)
        if form.is_valid():
            expense: Expense = form.save(commit=False)
            expense.created_by = request.user
            expense.save()
            return redirect("expenses:detail", slug=expense.slug)
        else:
            # Add field-specific error messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            messages.error(request, "Please correct the errors below.")
    else:
        form = ExpenseForm()
    # configure rest of template context
    intro = "Create Expense"
    breadcrumbs = [
        {"title": "Expense List", "url": reverse("expenses:list")},
        {"title": "Create Expense", "url": None},
    ]
    cancel_url = reverse("expenses:list")

    context = {
        "block_title": "Create Expense",
        "breadcrumbs": breadcrumbs,
        "title": "Create Expense",
        "intro": intro,
        "form": form,
        "submit_text": "Create",
        "cancel_url": cancel_url,
        "first_field": "item",
    }
    return render(request, "form/object.html", context)


@login_required
def list(request: HttpRequest) -> HttpResponse:
    # check if category is in parameters
    category_slug = request.GET.get("category")
    context = {
        "delete_modal_url": reverse("expenses:expense_delete_modal"),
        "expenses": Expense.objects.filter(is_deleted=False).order_by("-date"),
        "form": ExpenseForm(),
    }
    if category_slug:
        category = Category.objects.filter(slug=category_slug, is_deleted=False)
        if not category:
            return redirect("expenses:expense_list")
        context["category"] = category  # type: ignore[assignment]

    return render(request, "expenses/expense_list.html", context)


@login_required
def expense_edit(request, slug: str) -> HttpResponse:
    """Edit an existing expense"""
    expense = get_object_or_404(Expense, slug=slug, is_deleted=False)

    if request.method == "POST":
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.updated_by = request.user
            expense.save()
            messages.success(request, f"Expense '{expense.item}' updated successfully.")
            return redirect("expenses:detail", slug=expense.slug)
        else:
            # Add field-specific error messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            messages.error(request, "Please correct the errors below.")
    else:
        form = ExpenseForm(instance=expense)
    # configure rest of template context
    intro = f"Edit Expense: {expense.item}"
    breadcrumbs = [
        {"title": "Expense List", "url": reverse("expenses:list")},
        {"title": expense.category, "url": reverse("expenses:category_detail", args=[expense.category.slug])},
        {"title": expense, "url": None},
    ]
    cancel_url = reverse("expenses:detail", args=[expense.slug])

    context = {
        "block_title": "Edit Expense",
        "breadcrumbs": breadcrumbs,
        "title": f"Edit Expense: {expense.item}",
        "intro": intro,
        "form": form,
        "object": expense,
        "submit_text": "Save Changes",
        "cancel_url": cancel_url,
        "first_field": "item",
    }
    return render(request, "form/object.html", context)


@login_required
def expense_delete_modal(request: HttpRequest) -> HttpResponse:
    """
    Render the modal content for deleting an expense.
    """
    expense_slug = request.GET.get("slug")
    if not expense_slug:
        return JsonResponse({"error": "Expense slug is required"}, status=400)
    try:
        expense = Expense.objects.get(slug=expense_slug, is_deleted=False)
    except Expense.DoesNotExist:
        return JsonResponse({"error": "Expense not found"}, status=404)

    context = {
        "object": expense,
        "object_type": "Expense",
        "action_url": reverse("expenses:expense_delete", args=[expense.slug]),
    }
    return render(request, "components/modal/object_delete_modal_body.html", context)


@login_required
def expense_delete(request, slug: str) -> HttpResponse:
    """Delete an expense"""
    expense = get_object_or_404(Expense, slug=slug, is_deleted=False)

    if request.method == "POST":
        expense.is_deleted = True
        expense.save()
        messages.success(request, f"Expense '{expense.item}' deleted successfully.")
        return redirect("expenses:list")
    else:
        messages.warning(request, "Method not allowed. Please use POST to delete an expense.")
        return redirect("expenses:detail", slug=expense.slug)


@login_required
def expense_data(request) -> JsonResponse:
    """JSON endpoint for DataTables"""
    category_slug = request.GET.get("category")

    # Get DataTables parameters
    draw = int(request.GET.get("draw", 1))
    start = int(request.GET.get("start", 0))
    length = int(request.GET.get("length", 10))
    search_value = request.GET.get("search[value]", "")

    # Handle sorting
    order_column_index = request.GET.get("order[0][column]")
    order_direction = request.GET.get("order[0][dir]")
    order_column_name = request.GET.get(f"columns[{order_column_index}][name]")

    if category_slug:
        queryset = Expense.objects.filter(category__slug=category_slug, is_deleted=False).select_related("category")
    else:
        queryset = Expense.objects.filter(is_deleted=False).select_related("category")

    records_total = queryset.count()

    # Apply search filter if provided
    if search_value:
        queryset = queryset.filter(Q(item__icontains=search_value) | Q(description__icontains=search_value))
    records_filtered = queryset.count()

    # sort the queryset if order_column_name is provided
    if order_column_name:
        if order_direction == "desc":
            order_by = f"-{order_column_name}"
        else:
            order_by = order_column_name
        queryset = queryset.order_by(order_by)

    # Apply pagination
    queryset = queryset[start : start + length]

    # Prepare data for DataTables
    data = []

    for expense in queryset:
        detail_url = reverse("expenses:detail", args=[expense.slug])

        menu_content = format_row_action_cell(
            detail_url=detail_url,
            update_url=reverse("expenses:expense_edit", args=[expense.slug]),
            slug=expense.slug,
        )

        if expense.list_entries.exists():
            est_amount = float(expense.calculated_estimated_amount())
            actual_amount = float(expense.calculated_actual_amount())
            item = f'<a href="{detail_url}" class="link link-primary"><span class="italic">{expense.item}</span></a>'
        else:
            est_amount = float(expense.estimated_amount) if expense.estimated_amount else 0
            actual_amount = float(expense.actual_amount) if expense.actual_amount else 0
            item = f'<a href="{detail_url}" class="link link-primary">{expense.item}</a>'

        data.append(
            {
                "id": str(expense.id),
                "item": item,
                "category": expense.category.name if expense.category else "Uncategorized",
                "date": expense.date.strftime("%Y-%m-%d") if expense.date else None,
                "estimated_amount": est_amount,
                "actual_amount": actual_amount,
                "slug": expense.slug,
                "menu_content": menu_content,
            }
        )
    return JsonResponse(
        {
            "draw": draw,
            "recordsTotal": records_total,
            "recordsFiltered": records_filtered,
            "data": data,
        }
    )


@login_required
def expense_import(request) -> HttpResponse:
    """Import expenses from Excel file"""
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please select an Excel file to upload.")
            return redirect("expenses:summary")

        if not excel_file.name.endswith((".xlsx", ".xls")):
            messages.error(request, "Please upload a valid Excel file (.xlsx or .xls).")
            return redirect("expenses:summary")

        try:
            # Read the Excel file
            df = pl.read_excel(BytesIO(excel_file.read()))

            # Validate required columns
            required_columns = ["item"]
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
                return redirect("expenses:summary")

            # Import expenses
            created_count = 0
            updated_count = 0
            error_count = 0
            created_categories = 0

            for row in df.to_dicts():
                try:
                    item = str(row.get("item", "")).strip()
                    if not item or item.lower() == "nan":
                        error_count += 1
                        continue

                    # Handle category - create if doesn't exist
                    category = None
                    category_name = str(row.get("category", "")).strip()
                    if category_name and category_name.lower() != "nan":
                        category, created = Category.objects.get_or_create(
                            name=category_name,
                            defaults={
                                "slug": slugify(category_name),
                                "created_by": request.user,
                                "is_deleted": False,
                            },
                        )
                        if created:
                            created_categories += 1

                    # Parse date
                    expense_date = None
                    date_str = str(row.get("date", "")).strip()
                    if date_str and date_str.lower() != "nan":
                        try:
                            expense_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                        except ValueError:
                            try:
                                expense_date = datetime.strptime(date_str, "%m/%d/%Y").date()
                            except ValueError:
                                expense_date = None

                    # Parse amounts
                    estimated_amount = None
                    actual_amount = None
                    quantity = 1

                    estimated_str = str(row.get("estimated_amount", "")).strip()
                    if estimated_str and estimated_str.lower() != "nan":
                        try:
                            estimated_amount = Decimal(str(estimated_str).replace("$", "").replace(",", ""))
                        except (InvalidOperation, ValueError):
                            estimated_amount = None

                    actual_str = str(row.get("actual_amount", "")).strip()
                    if actual_str and actual_str.lower() != "nan":
                        try:
                            actual_amount = Decimal(str(actual_str).replace("$", "").replace(",", ""))
                        except (InvalidOperation, ValueError):
                            actual_amount = None

                    quantity_str = str(row.get("quantity", "")).strip()
                    if quantity_str and quantity_str.lower() != "nan":
                        try:
                            quantity = int(float(quantity_str))
                        except (ValueError, TypeError):
                            quantity = 1

                    # Check if expense already exists
                    existing_expense = None
                    if category:
                        existing_expense = Expense.objects.filter(
                            item__iexact=item, category=category, is_deleted=False
                        ).first()
                    else:
                        existing_expense = Expense.objects.filter(
                            item__iexact=item, category__isnull=True, is_deleted=False
                        ).first()

                    expense_data = {
                        "item": item,
                        "description": str(row.get("description", "")).strip()
                        if row.get("description", None) is not None
                        else "",
                        "date": expense_date,
                        "category": category,
                        "quantity": quantity,
                        "estimated_amount": estimated_amount,
                        "actual_amount": actual_amount,
                        "url": str(row.get("url", "")).strip() if row.get("url", None) is not None else "",
                    }

                    # Clean empty strings
                    for key, value in expense_data.items():
                        if isinstance(value, str) and (value.lower() == "nan" or not value):
                            expense_data[key] = "" if key not in ["item"] else expense_data[key]

                    if existing_expense:
                        # Update existing expense
                        for key, value in expense_data.items():
                            if value is not None and value != "":  # Only update non-empty values
                                setattr(existing_expense, key, value)
                        existing_expense.updated_by = request.user
                        existing_expense.save()
                        updated_count += 1
                    else:
                        # Create new expense
                        expense = Expense(**expense_data)
                        expense.created_by = request.user
                        expense.save()
                        created_count += 1

                except Exception as e:
                    logger.error(e)
                    error_count += 1
                    continue

            # Success message
            message_parts = []
            if created_count > 0:
                message_parts.append(f"{created_count} expense{'s' if created_count != 1 else ''} created")
            if updated_count > 0:
                message_parts.append(f"{updated_count} expense{'s' if updated_count != 1 else ''} updated")
            if created_categories > 0:
                message_parts.append(
                    f"{created_categories} new categor{'ies' if created_categories != 1 else 'y'} created"
                )

            if message_parts:
                messages.success(request, f"Import completed: {', '.join(message_parts)}.")

            if error_count > 0:
                messages.warning(request, f"{error_count} row{'s' if error_count != 1 else ''} skipped due to errors.")

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")

    return redirect("expenses:summary")


@login_required
def detail(request, slug: str) -> HttpResponse:
    """View for expense detail"""
    try:
        expense = Expense.objects.get(slug=slug, is_deleted=False)
    except Expense.DoesNotExist:
        messages.error(request, "Expense not found.")
        return redirect("expenses:summary")

    breadcrumbs = [
        {"title": "Budget Summary", "url": reverse("expenses:summary")},
        {"title": expense.category, "url": reverse("expenses:category_detail", args=[expense.category.slug])},
        {"title": f"{expense}", "url": None},
    ]
    attach_url_query = QueryDict("", mutable=True)
    attach_url_query["next"] = request.path
    context = {
        "block_title": f"{expense}",
        "breadcrumbs": breadcrumbs,
        "title": f"{expense}",
        "object": expense,
        "edit_url": reverse("expenses:expense_edit", args=[expense.slug]),
        "status": expense.purchased,
        "status_text": "Purchased" if expense.purchased else "Pending",
        "link_url": expense.url,
        "image_url": None,
        "delete_modal_url": reverse("expenses:expense_delete_modal"),
        "list_entries": expense.list_entries.filter(is_deleted=False),
        "attachments": Attachment.objects.attachments_for_object(expense).all(),
        "attach_form": AttachmentUploadForm(),
        "attach_submit_url": str(
            reverse(
                "attachments:add_attachment",
                kwargs={"app_label": "expenses", "model_name": "Expense", "pk": expense.pk},
                query=attach_url_query,
            )
        ),
    }

    return render(request, "expenses/expense_detail.html", context)


@login_required
def template_download(request) -> HttpResponse:
    """Download Excel template for expense import"""
    # Create a workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses Template"

    # Define headers
    headers = ["item", "description", "date", "category", "quantity", "estimated_amount", "actual_amount", "url"]

    # Add headers to the first row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)

    # Add sample data
    sample_data = [
        [
            "Wedding Photography",
            "Professional photographer for ceremony and reception",
            "2024-08-15",
            "Photography",
            1,
            2500.00,
            2800.00,
            "https://example.com/wedding-photography",
        ],
        [
            "Wedding Cake",
            "3-tier vanilla cake with custom decorations",
            "2024-08-15",
            "Catering",
            1,
            800.00,
            750.00,
            "https://example.com/wedding-cake",
        ],
        [
            "Bridal Dress",
            "Designer wedding dress with alterations",
            "2024-06-01",
            "Attire",
            1,
            1500.00,
            1650.00,
            "https://example.com/bridal-dress",
        ],
        [
            "Flower Arrangements",
            "Bridal bouquet and centerpieces",
            "2024-08-15",
            "Flowers",
            8,
            150.00,
            140.00,
            "https://example.com/flower-arrangements",
        ],
        [
            "Wedding Venue",
            "Reception hall rental for 6 hours",
            "2024-08-15",
            "Venue",
            1,
            3000.00,
            "",
            "https://example.com/wedding-venue",
        ],
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = HttpResponse(
        output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="expenses_template.xlsx"'

    return response


@login_required
def category_detail(request, slug: str) -> HttpResponse:
    try:
        category = Category.objects.get(slug=slug, is_deleted=False)
        expenses = Expense.objects.filter(category=category, is_deleted=False).order_by("-date")
    except Category.DoesNotExist:
        messages.error(request, "Category not found.")
        return redirect("expenses:expense_summary")

    breadcrumbs = [
        {"title": "Budget Summary", "url": reverse("expenses:summary")},
        {"title": f"{category}", "url": None},
    ]

    estimated = expenses.aggregate(total=Sum("estimated_amount", filter=Q(estimated_amount__isnull=False)))[
        "total"
    ] or Decimal("0.00")
    actual = expenses.aggregate(total=Sum("actual_amount", filter=Q(actual_amount__isnull=False)))["total"] or Decimal(
        "0.00"
    )
    context = {
        "block_title": f"{category}",
        "breadcrumbs": breadcrumbs,
        "title": f"{category}",
        "object": category,
        "edit_url": reverse("expenses:category_edit", args=[category.slug]),
        # "status": category.purchased,
        # "status_text": "Purchased" if category.purchased else "Pending",
        "link_url": None,
        "image_url": None,
        "delete_modal_url": reverse("expenses:category_delete_modal"),
        "expenses": expenses,
        "total_estimated": estimated,
        "total_actual": actual,
    }

    return render(request, "expenses/category_detail.html", context)


@login_required
def category_create(request) -> HttpResponse:
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            category: Category = form.save(commit=False)
            category.created_by = request.user
            category.save()
            return redirect("expenses:summary")
    else:
        form = CategoryForm()
    intro = f"Create Category"
    breadcrumbs = [
        {"title": "Budget", "url": reverse("expenses:summary")},
        {"title": "Create", "url": None},
    ]
    cancel_url = reverse("expenses:summary")

    context = {
        "block_title": "Create Category",
        "breadcrumbs": breadcrumbs,
        "title": f"Create Category",
        "intro": intro,
        "form": form,
        "submit_text": "Save Changes",
        "cancel_url": cancel_url,
        "first_field": "name",
    }
    return render(request, "form/object.html", context)


@login_required
def category_edit(request, slug: str) -> HttpResponse:
    category = Category.objects.get(slug=slug, is_deleted=False)
    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            category = form.save(commit=False)
            category.updated_by = request.user
            category.save()
            return redirect("expenses:category_detail", slug=category.slug)
    else:
        form = CategoryForm(instance=category)
    intro = f"Edit Category: {category}"
    breadcrumbs = [
        {"title": "Budget", "url": reverse("expenses:summary")},
        {"title": category, "url": reverse("expenses:category_detail", args=[category.slug])},
        {"title": "Edit", "url": None},
    ]
    cancel_url = reverse("expenses:category_detail", args=[category.slug])

    context = {
        "block_title": "Edit Category",
        "breadcrumbs": breadcrumbs,
        "title": f"Edit Category: {category.name}",
        "intro": intro,
        "form": form,
        "object": category,
        "submit_text": "Save Changes",
        "cancel_url": cancel_url,
        "first_field": "name",
    }
    return render(request, "form/object.html", context)


@login_required
def category_delete(request, slug: str) -> HttpResponse:
    category = get_object_or_404(Category, slug=slug, is_deleted=False)
    uncategorized = get_object_or_404(Category, name="Uncategorized", is_deleted=False)
    if request.method == "POST":
        category.is_deleted = True
        category.save()
        # Move expenses to Uncategorized
        expenses = Expense.objects.filter(category=category)
        for expense in expenses:
            expense.category = uncategorized
            expense.save()
    return redirect("expenses:summary")


@login_required
def category_delete_modal(request: HttpRequest) -> HttpResponse:
    """
    Render the modal content for deleting a category.
    """
    category_slug = request.GET.get("slug")
    if not category_slug:
        return JsonResponse({"error": "Category slug is required"}, status=400)
    try:
        category = Category.objects.get(slug=category_slug, is_deleted=False)
    except Category.DoesNotExist:
        return JsonResponse({"error": "Category not found"}, status=404)

    context = {
        "action_url": reverse("expenses:category_delete", args=[category.slug]),
        "custom_message": f"This action cannot be undone. All expenses associated to this '{category}' Category will be moved to the 'Uncategorized' category.",
    }
    return render(request, "components/modal/object_delete_modal_body.html", context)
