import io
from click import File
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse, QueryDict
from django.urls import reverse
import polars as pl
import openpyxl
from io import BytesIO
import logging

from attachments.forms import AttachmentUploadForm
from attachments.models import Attachment
from list.models import ListEntry
from .models import Contact
from .forms import ContactForm

logger = logging.getLogger(__name__)


@login_required
def contact_list(request):
    """List all contacts with search and pagination"""
    search_query = request.GET.get("search", "")
    contacts = Contact.objects.filter(is_deleted=False)

    if search_query:
        contacts = contacts.filter(
            Q(name__icontains=search_query)
            | Q(company__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(phone__icontains=search_query)
            | Q(website__icontains=search_query)
            | Q(category__icontains=search_query)
        )

    contacts = contacts.order_by("name")

    paginator = Paginator(contacts, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "search_query": search_query,
        "total_contacts": contacts.count(),
        "delete_modal_url": reverse("contacts:contact_delete_modal"),
    }

    return render(request, "contacts/contact_summary.html", context)


@login_required
def contact_detail(request, slug):
    """Show contact details"""
    try:
        contact = Contact.objects.get(slug=slug, is_deleted=False)
    except Contact.DoesNotExist:
        messages.error(request, "Contact not found.")
        return redirect("contacts:list")

    # get attachments from expense objects where this contact is the vendor
    attachments = Attachment.objects.filter(expense__vendor=contact)

    breadcrumbs = [
        {"title": "Contacts", "url": reverse("contacts:list")},
        {"title": f"{contact}", "url": None},
    ]
    attach_url_query = QueryDict("", mutable=True)
    attach_url_query["next"] = request.path
    context = {
        "block_title": f"{contact}",
        "breadcrumbs": breadcrumbs,
        "title": f"{contact}",
        "object": contact,
        "edit_url": reverse("contacts:contact_edit", args=[contact.slug]),
        "status": None,
        "status_text": None,
        "link_url": contact.website,
        "image_url": None,
        "delete_modal_url": reverse("contacts:contact_delete_modal"),
        "shopping_list": ListEntry.objects.filter(vendor=contact, purchased=False, is_deleted=False).order_by("item"),
        "attachments": Attachment.objects.attachments_for_object(contact).all(),
        "attach_form": AttachmentUploadForm(),
        "attach_submit_url": str(
            reverse(
                "attachments:add_attachment",
                kwargs={"app_label": "contacts", "model_name": "Contact", "pk": contact.pk},
                query=attach_url_query,
            )
        ),
    }

    return render(request, "contacts/contact_detail.html", context)


@login_required
def contact_create(request):
    """Create a new contact"""
    if request.method == "POST":
        form = ContactForm(request.POST)
        if form.is_valid():
            contact: Contact = form.save(commit=False)
            contact.created_by = request.user
            contact.save()
            messages.success(request, f'Contact "{contact.name}" was created successfully.')
            return redirect("contacts:detail", slug=contact.slug)
    else:
        form = ContactForm()

    intro = f"Create New Contact"
    breadcrumbs = [
        {"title": "Contacts", "url": reverse("contacts:list")},
        {"title": "Create New Contact", "url": None},
    ]
    cancel_url = reverse("contacts:list")

    context = {
        "block_title": "Create New Contact",
        "breadcrumbs": breadcrumbs,
        "title": "Create New Contact",
        "intro": intro,
        "form": form,
        "object": Contact,
        "submit_text": "Create",
        "cancel_url": cancel_url,
        "first_field": "name",
        "file": False,
    }
    return render(request, "shared_helpers/form/object.html", context)


@login_required
def contact_edit(request, slug):
    """Update an existing contact"""
    contact = get_object_or_404(Contact, slug=slug, is_deleted=False)

    if request.method == "POST":
        form = ContactForm(request.POST, instance=contact)
        if form.is_valid():
            contact: Contact = form.save(commit=False)
            contact.updated_by = request.user
            contact.save()
            messages.success(request, f'Contact "{contact.name}" was updated successfully.')
            return redirect("contacts:detail", slug=contact.slug)
        else:
            # Add field-specific error messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            messages.error(request, "Please correct the errors below.")
    else:
        form = ContactForm(instance=contact)

    intro = f"Update {contact.name}"
    breadcrumbs = [
        {"title": "Contacts", "url": reverse("contacts:list")},
        {"title": contact.name, "url": reverse("contacts:detail", args=[contact.slug])},
        {"title": "Update", "url": None},
    ]
    cancel_url = reverse("contacts:detail", args=[contact.slug])

    context = {
        "block_title": "Edit Contact",
        "breadcrumbs": breadcrumbs,
        "title": f"Edit Contact: {contact.name}",
        "intro": intro,
        "form": form,
        "object": Contact,
        "submit_text": "Save Changes",
        "cancel_url": cancel_url,
        "first_field": "name",
        "file": False,
        "hugerte_enabled": True,
        "selector": "id_notes",
    }
    return render(request, "shared_helpers/form/object.html", context)


@login_required
def contact_delete_modal(request):
    slug = request.GET.get("slug")
    if not slug:
        return JsonResponse({"error": "Contact slug is required"}, status=400)
    try:
        contact = Contact.objects.get(slug=slug, is_deleted=False)
    except Contact.DoesNotExist:
        return JsonResponse({"error": "Contact not found"}, status=404)

    context = {
        "object": contact,
        "object_type": "Contact",
        "action_url": reverse("contacts:contact_delete", args=[contact.slug]),
    }
    return render(request, "shared_helpers/modal/object_delete_body.html", context)


@login_required
def contact_delete(request, slug):
    if request.method == "POST":
        try:
            contact_obj = Contact.objects.get(slug=slug, is_deleted=False)
        except Contact.DoesNotExist:
            messages.error(request, "Contact not found.")
            return redirect("contacts:list")
        contact_obj.is_deleted = True
        contact_obj.save()
        messages.success(request, "Contact deleted successfully.")
        return redirect(f"{reverse('contacts:list')}")
    else:
        messages.error(request, "Invalid request method.")
        return redirect("contacts:list")


@login_required
def contact_import(request):
    """Import contacts from Excel file"""
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please select an Excel file to upload.")
            return redirect("contacts:list")

        if not excel_file.name.endswith((".xlsx", ".xls")):
            messages.error(request, "Please upload a valid Excel file (.xlsx or .xls).")
            return redirect("contacts:list")

        try:
            # Read the Excel file
            df = pl.read_excel(BytesIO(excel_file.read()))

            # Validate required columns
            required_columns = ["name"]
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
                return redirect("contacts:list")

            # Import contacts
            created_count = 0
            updated_count = 0
            error_count = 0

            for row in df.to_dicts():
                try:
                    name = str(row.get("name", "")).strip() if row.get("name", None) else None
                    company = str(row.get("company", "")).strip() if row.get("company", None) else None

                    # Check if contact already exists
                    existing_contact = Contact.objects.filter(
                        name__iexact=name, company__iexact=company, is_deleted=False
                    ).first()

                    contact_data = {
                        "name": name,
                        "company": company,
                        "email": str(row.get("email", "")).strip() if row.get("email", None) is not None else "",
                        "phone": str(row.get("phone", "")).strip() if row.get("phone", None) is not None else "",
                        "website": str(row.get("website", "")).strip() if row.get("website", None) is not None else "",
                        "category": str(row.get("category", "")).strip()
                        if row.get("category", None) is not None
                        else "",
                        "notes": str(row.get("notes", "")).strip() if row.get("notes", None) is not None else "",
                    }

                    if existing_contact:
                        # Update existing contact
                        for key, value in contact_data.items():
                            if value is not None and value != "":  # Only update non-empty values
                                setattr(existing_contact, key, value)
                        existing_contact.updated_by = request.user
                        existing_contact.save()
                        updated_count += 1
                    else:
                        # Create new contact
                        contact = Contact(**contact_data)
                        contact.created_by = request.user
                        contact.save()
                        created_count += 1

                except Exception as e:
                    messages.error(request, f"Error processing contact row {row}: {str(e)}")
                    error_count += 1
                    continue

            # Success message
            message_parts = []
            if created_count > 0:
                message_parts.append(f"{created_count} contact{'s' if created_count != 1 else ''} created")
            if updated_count > 0:
                message_parts.append(f"{updated_count} contact{'s' if updated_count != 1 else ''} updated")

            if message_parts:
                messages.success(request, f"Import completed: {', '.join(message_parts)}.")

            if error_count > 0:
                messages.warning(request, f"{error_count} row{'s' if error_count != 1 else ''} skipped due to errors.")

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")

    return redirect("contacts:list")


@login_required
def template_download(request):
    """Download Excel template for contact import"""
    # Create a workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts Template"

    # Define headers
    headers = ["name", "company", "email", "phone", "website", "category", "notes"]

    # Add headers to the first row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)

    # Add sample data
    sample_data = [
        [
            "John Doe",
            "ABC Photography",
            "john@abcphoto.com",
            "555-0123",
            "https://abcphoto.com",
            "Vendor",
            "Wedding photographer",
        ],
        [
            "Jane Smith",
            "Elegant Flowers",
            "jane@elegantflowers.com",
            "555-0456",
            "https://elegantflowers.com",
            "Vendor",
            "Florist for ceremony and reception",
        ],
        [
            "Mike Johnson",
            "Sound & Light Co",
            "mike@soundlight.com",
            "555-0789",
            "https://soundlight.com",
            "Vendor",
            "DJ and lighting services",
        ],
        ["Sarah Wilson", "", "sarah@email.com", "555-0321", "", "Guest", "College friend"],
        [
            "The Grand Ballroom",
            "",
            "events@grandballroom.com",
            "555-0654",
            "https://grandballroom.com",
            "Venue",
            "Reception venue option",
        ],
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = HttpResponse(
        output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="contacts_template.xlsx"'

    return response
